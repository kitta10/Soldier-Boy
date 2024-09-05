import openpyxl

# usernames=[("electricalex0808@gmail.com","selenium"),("chotabheem69@gmail.com","chintapakdum"),("amayudayan@gmail.com","Sonu3110")]
# login_creds.cell(1,1)

def get_list_from_excel(excel_name,sheet_name):
    excel = openpyxl.load_workbook(excel_name)
    sheet = excel[sheet_name]
    credentials = []
    for row in range(1, sheet.max_row + 1):
        nested_creds = []
        for column in range(1, sheet.max_column + 1):
            data = sheet.cell(row, column)
            nested_creds.append(data.value)
        credentials.append(nested_creds)
    return credentials

def add_value_to_excel(excel_name,sheet_name):
    excel=openpyxl.load_workbook(excel_name)
    sheet=excel[sheet_name]
    sheet.cell(10,1).value="new value"
    excel.save(excel_name)

add_value_to_excel("SeleniumProject.xlsx","login_creds")
